"""Device management router - brands, gadget types, device catalog."""
from __future__ import annotations

import io
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from veriqko.db.base import get_db
from veriqko.dependencies import get_current_user
from veriqko.devices.models import Brand, Device, GadgetType
from veriqko.devices.schemas import (
    BrandCreate,
    BrandResponse,
    BrandUpdate,
    DeviceCreate,
    DeviceResponse,
    DeviceUpdate,
    GadgetTypeCreate,
    GadgetTypeResponse,
    GadgetTypeUpdate,
)
from veriqko.users.models import User

router = APIRouter(prefix="/admin", tags=["devices"])

# ─────────────────────────── Brands ────────────────────────────

@router.get("/brands", response_model=list[BrandResponse])
async def list_brands(db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    result = await db.execute(select(Brand).order_by(Brand.name))
    return result.scalars().all()

@router.post("/brands", response_model=BrandResponse, status_code=status.HTTP_201_CREATED)
async def create_brand(brand_in: BrandCreate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    brand = Brand(**brand_in.model_dump())
    db.add(brand)
    await db.commit()
    await db.refresh(brand)
    return brand

@router.put("/brands/{brand_id}", response_model=BrandResponse)
async def update_brand(brand_id: str, brand_in: BrandUpdate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    brand = await db.get(Brand, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    for field, value in brand_in.model_dump(exclude_unset=True).items():
        setattr(brand, field, value)
    await db.commit()
    await db.refresh(brand)
    return brand

@router.delete("/brands/{brand_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_brand(brand_id: str, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    brand = await db.get(Brand, brand_id)
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    await db.delete(brand)
    await db.commit()

# ─────────────────────────── Gadget Types ──────────────────────

@router.get("/gadget-types", response_model=list[GadgetTypeResponse])
async def list_gadget_types(db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    result = await db.execute(select(GadgetType).order_by(GadgetType.name))
    return result.scalars().all()

@router.post("/gadget-types", response_model=GadgetTypeResponse, status_code=status.HTTP_201_CREATED)
async def create_gadget_type(type_in: GadgetTypeCreate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    gadget_type = GadgetType(**type_in.model_dump())
    db.add(gadget_type)
    await db.commit()
    await db.refresh(gadget_type)
    return gadget_type

@router.put("/gadget-types/{type_id}", response_model=GadgetTypeResponse)
async def update_gadget_type(type_id: str, type_in: GadgetTypeUpdate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    gadget_type = await db.get(GadgetType, type_id)
    if not gadget_type:
        raise HTTPException(status_code=404, detail="Gadget type not found")
    for field, value in type_in.model_dump(exclude_unset=True).items():
        setattr(gadget_type, field, value)
    await db.commit()
    await db.refresh(gadget_type)
    return gadget_type

@router.delete("/gadget-types/{type_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_gadget_type(type_id: str, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    gadget_type = await db.get(GadgetType, type_id)
    if not gadget_type:
        raise HTTPException(status_code=404, detail="Gadget type not found")
    await db.delete(gadget_type)
    await db.commit()

# ─────────────────────────── Devices ──────────────────────────

def _device_query():
    return select(Device).options(joinedload(Device.brand), joinedload(Device.gadget_type))

@router.get("/devices", response_model=list[DeviceResponse])
async def list_devices(db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    result = await db.execute(_device_query().order_by(Device.model))
    return result.scalars().all()

@router.post("/devices", response_model=DeviceResponse, status_code=status.HTTP_201_CREATED)
async def create_device(device_in: DeviceCreate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    device = Device(**device_in.model_dump())
    db.add(device)
    await db.commit()
    result = await db.execute(_device_query().where(Device.id == device.id))
    return result.scalar_one()

@router.get("/devices/{device_id}", response_model=DeviceResponse)
async def get_device(device_id: str, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    result = await db.execute(_device_query().where(Device.id == device_id))
    device = result.scalar_one_or_none()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    return device

@router.put("/devices/{device_id}", response_model=DeviceResponse)
async def update_device(device_id: str, device_in: DeviceUpdate, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    for field, value in device_in.model_dump(exclude_unset=True).items():
        setattr(device, field, value)
    await db.commit()
    result = await db.execute(_device_query().where(Device.id == device.id))
    return result.scalar_one()

@router.delete("/devices/{device_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_device(device_id: str, db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    device = await db.get(Device, device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    await db.delete(device)
    await db.commit()

# ─────────────────────────── Excel Import/Export ──────────────

@router.get("/devices/export/excel")
async def export_devices_excel(db: Annotated[AsyncSession, Depends(get_db)], current_user: Annotated[User, Depends(get_current_user)]):
    """Export the entire device catalog to an Excel file."""
    import openpyxl

    result = await db.execute(_device_query().order_by(Device.model))
    devices = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Catalog"

    headers = ["Brand", "Type", "Model", "Model Number", "Colour", "Storage", "ID"]
    ws.append(headers)

    for d in devices:
        ws.append([
            d.brand.name if d.brand else "",
            d.gadget_type.name if d.gadget_type else "",
            d.model,
            d.model_number or "",
            d.colour or "",
            d.storage or "",
            d.id,
        ])

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=device_catalog.xlsx"},
    )


@router.post("/devices/import/excel", status_code=status.HTTP_200_OK)
async def import_devices_excel(
    file: Annotated[UploadFile, File(description="Excel file (.xlsx) with columns: Brand, Type, Model, Model Number, Colour, Storage")],
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    """Import devices from an Excel file. New rows are added; rows with matching ID are skipped."""
    import openpyxl

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    # Load lookup maps
    brand_result = await db.execute(select(Brand))
    brand_map = {b.name.lower(): b for b in brand_result.scalars().all()}

    type_result = await db.execute(select(GadgetType))
    type_map = {t.name.lower(): t for t in type_result.scalars().all()}

    created = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row_num, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        brand_name, type_name, model, model_number, colour, storage = (row + (None,) * 6)[:6]

        # Find or create Brand
        brand = brand_map.get(str(brand_name).lower() if brand_name else "")
        if not brand and brand_name:
            brand = Brand(name=str(brand_name))
            db.add(brand)
            await db.flush()
            brand_map[brand.name.lower()] = brand

        if not brand:
            errors.append(f"Row {row_num}: missing brand")
            continue

        # Find or create GadgetType
        gtype = type_map.get(str(type_name).lower() if type_name else "")
        if not gtype and type_name:
            gtype = GadgetType(name=str(type_name))
            db.add(gtype)
            await db.flush()
            type_map[gtype.name.lower()] = gtype

        if not gtype:
            errors.append(f"Row {row_num}: missing type")
            continue

        if not model:
            errors.append(f"Row {row_num}: missing model")
            continue

        device = Device(
            brand_id=brand.id,
            type_id=gtype.id,
            model=str(model),
            model_number=str(model_number) if model_number else None,
            colour=str(colour) if colour else None,
            storage=str(storage) if storage else None,
        )
        db.add(device)
        created += 1

    await db.commit()
    return {"created": created, "errors": errors}
